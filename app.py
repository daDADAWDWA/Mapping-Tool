"""
Stack View Generator -- Streamlit UI

Run locally with:
    streamlit run app.py

Workflow:
    1. Pick a project (or add a new one by uploading its template once)
    2. Upload this run's Transaction CSV
    3. Upload this run's Inventory Excel
    4. Generate -> preview summary -> download the filled .xlsx
"""

import os
import shutil
import tempfile
from pathlib import Path

import streamlit as st

from core.config import describe_sources, list_models, load_settings, masked, save_model
from core.pipeline import generate_stack_view

PROJECTS_DIR = Path(__file__).parent / "projects"
PROJECTS_DIR.mkdir(exist_ok=True)

st.set_page_config(page_title="Stack View Generator", layout="centered")
st.title("🏢 Stack View Generator")
st.caption("Turns a Transaction CSV + Inventory Excel into a filled-in Stack View, for any project template.")


def list_projects():
    found = []
    for p in PROJECTS_DIR.iterdir():
        if not p.is_dir():
            continue
        has_own_template = (p / "template.csv").exists()
        has_tower_subdirs = any(
            sub.is_dir() and (sub / "template.csv").exists() for sub in p.iterdir()
        )
        if has_own_template or has_tower_subdirs:
            found.append(p.name)
    return sorted(found)


def project_towers(project_name):
    """List of tower letters already set up for this project, or [] if it's single-tower."""
    proj_dir = PROJECTS_DIR / project_name
    return sorted([
        sub.name.upper() for sub in proj_dir.iterdir()
        if sub.is_dir() and (sub / "template.csv").exists()
    ])


# ---------------------------------------------------------------------------
# Sidebar: project management
# ---------------------------------------------------------------------------
with st.sidebar:
    st.header("Projects")
    projects = list_projects()
    st.write(f"{len(projects)} project(s) available.")

    with st.expander("➕ Add a new project or tower"):
        st.caption(
            "Leave 'Tower letter' blank for a simple single-tower project. "
            "Fill it in (A, B, C...) to add that tower to a project -- reuse "
            "the same project name to add more towers to it later."
        )
        new_name = st.text_input("Project name", key="new_project_name")
        new_tower = st.text_input("Tower letter (optional)", key="new_tower_letter", max_chars=3)
        new_template = st.file_uploader("Template CSV for this tower/project", type=["csv"], key="new_template")
        if st.button("Save"):
            if not new_name.strip():
                st.error("Give the project a name.")
            elif not new_template:
                st.error("Upload the template CSV first.")
            else:
                proj_dir = PROJECTS_DIR / new_name.strip()
                if new_tower.strip():
                    target_dir = proj_dir / new_tower.strip().upper()
                else:
                    target_dir = proj_dir
                target_dir.mkdir(parents=True, exist_ok=True)
                with open(target_dir / "template.csv", "wb") as f:
                    f.write(new_template.getbuffer())
                label = f"'{new_name}' (Tower {new_tower.strip().upper()})" if new_tower.strip() else f"'{new_name}'"
                st.success(f"Saved {label}. Select it above.")
                st.rerun()

    st.divider()
    st.subheader("AI assist (optional)")

    # The API key is NEVER entered here -- it lives in config.json or .env
    # inside the app folder. This block only reports whether one was found,
    # showing the last four characters so you can tell which key is loaded
    # without putting the key on screen.
    settings = load_settings()
    api_key = settings["api_key"]
    if api_key:
        st.success(
            f"API key loaded ({masked(api_key)}) from {settings['api_key_source']}."
        )
    else:
        st.info(
            "No API key found, so the switches below do nothing and the app runs "
            "entirely on pattern matching. To enable AI, put your key in "
            f"`config.json` next to `app.py`:\n\n"
            '```json\n{\n  "api_key": "sk-ant-...",\n  "model": '
            f'"{settings["model"]}"\n}}\n```\n'
            "or set `ANTHROPIC_API_KEY` in a `.env` file there. Restart the app "
            "after editing. Keep both files out of git.\n\n"
            "**Hosted on Streamlit Cloud?** Put it in the app's Secrets instead "
            "(Manage app → Settings → Secrets):\n\n"
            "```toml\nANTHROPIC_API_KEY = \"sk-ant-...\"\n```"
        )
        with st.expander("Where the key was looked for"):
            for label, found in describe_sources():
                st.write(("✅ " if found else "—  ") + label)
            st.caption(
                "Checked in that order, first hit wins. If a source shows a tick but "
                "the key still fails, it's likely malformed — check for stray quotes."
            )
    use_ai = st.checkbox(
        "Read the area out of hard transaction descriptions",
        value=bool(api_key),
        help="Only for rows where pattern matching found no carpet area in the "
             "Marathi/English description text, or couldn't tell which area "
             "type it was.",
    )
    use_ai_inventory = st.checkbox(
        "Match column names, parse the inventory PDF, decode unit numbers",
        value=bool(api_key),
        help="Finds the right column when your file uses an unfamiliar name (e.g. "
             "'Unit' instead of 'Unit No'); checks which PDF column is really the carpet area before any value "
             "is read; decodes odd unit numbers (G-04, PH-2, 1601/1602) into "
             "floor and series; and reconstructs the table if no rows could be "
             "detected at all. AI never supplies an area number when the PDF's "
             "own text can -- see the README.",
    )
    use_ai_documents = st.checkbox(
        "Read the brochure and agreements (Sections 3 & 4)",
        value=bool(api_key),
        help="Brochure PDF -> Section 3, agreement PDFs or photos -> Section 4. "
             "Neither is a table, so these can ONLY be read with AI -- turning "
             "this off means an uploaded brochure/agreement is ignored. Pages are "
             "keyword-shortlisted first so a long PDF doesn't get sent whole.",
    )
    # ---- model choice, per run ----
    st.divider()
    st.subheader("Model")
    model = settings["model"]
    if api_key:
        available = list_models(api_key)
        if available:
            options = available + ["Other (type it in)"]
            default_idx = available.index(model) if model in available else 0
            picked = st.selectbox(
                "Model for this run", options, index=default_idx,
                help="Fetched live from your account, so it's never a stale hardcoded "
                     "list. Applies to every AI call in this run.",
            )
            model = st.text_input("Model ID", value=model) \
                if picked == "Other (type it in)" else picked
        else:
            model = st.text_input(
                "Model ID", value=model,
                help="Couldn't fetch the model list from the API, so type the ID.",
            )
            st.caption("Model list unavailable — check the key if this is unexpected.")
        st.caption(f"Default is `{settings['model']}` (from {settings['model_source']}).")
        if model != settings["model"] and st.button("Make this the default"):
            if save_model(model):
                st.success(f"`{model}` saved to config.json as the default.")
                st.rerun()
            else:
                st.error("Couldn't write config.json — check the folder is writable.")
    else:
        st.caption(f"Would use `{model}` (from {settings['model_source']}) once a key is set.")

# ---------------------------------------------------------------------------
# Main flow
# ---------------------------------------------------------------------------
if not projects:
    st.warning("No projects yet. Add one from the sidebar (upload a template CSV) to get started.")
    st.stop()

project = st.selectbox("1. Choose a project", projects)

towers = project_towers(project)
if towers:
    st.caption(f"This project has {len(towers)} tower(s) set up: {', '.join(towers)}. "
               f"The output will have one sheet per tower.")

col1, col2 = st.columns(2)
with col1:
    txn_file = st.file_uploader("2. Transaction CSV", type=["csv"], key="txn")
with col2:
    inv_file = st.file_uploader("3. Inventory (Excel or PDF)", type=["xlsx", "xls", "pdf"], key="inv")

st.caption(
    "Optional evidence documents. The brochure fills Section 3, agreements fill "
    "Section 4, and both feed the Final Output priority chain: "
    "**agreement → CRE transaction → builder inventory → brochure**. "
    "Both need an API key, since neither document is a table."
)
col3, col4 = st.columns(2)
with col3:
    brochure_file = st.file_uploader("4. Builder brochure (optional)", type=["pdf"], key="broch")
with col4:
    agreement_files = st.file_uploader(
        "5. Agreements — PDF, multi-page TIFF or photos (optional, several at once)",
        type=["pdf", "png", "jpg", "jpeg", "webp", "heic", "heif", "tif", "tiff"],
        accept_multiple_files=True, key="agmts",
        help="A 200-page agreement is fine: pages are searched in priority order "
             "(Second Schedule first, then any page naming RERA Carpet / MOFA / "
             "Built-up, then floor plans) and the search stops at the first page that "
             "answers. A photo of just the schedule page is still the cheapest option. "
             "Name the file with the unit number (e.g. A_1201.jpg) as a fallback "
             "in case the photo crops the unit number out.",
    )

if brochure_file or agreement_files:
    if not api_key:
        st.warning("A brochure/agreement was attached but there's no API key, so they can't be read.")
    elif not use_ai_documents:
        st.warning(
            "A brochure/agreement was attached but \"Read the brochure and agreements\" "
            "is unticked in the sidebar, so they'll be ignored. These documents can only "
            "be read with AI."
        )

tolerance_ft = st.number_input(
    "Area tolerance (sq ft) — differences smaller than this count as the same area",
    min_value=0.0, max_value=50.0, value=5.0, step=1.0,
    help="Measurement noise, not a real difference. At 5, a 2 sq ft gap merges into "
         "the same Final Output row; a 20 sq ft gap splits off its own row and is flagged.",
)

st.divider()
st.caption(
    "Every property description is read automatically when you generate — unit "
    "number, tower, floor, series and every area with its type. The numbering "
    "convention is detected from the data, not assumed."
)

generate = st.button("6. Generate Stack View", type="primary", disabled=not (txn_file and inv_file))

if generate:
    with st.spinner("Processing..."):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            txn_path = tmp / "transaction.csv"
            inv_ext = Path(inv_file.name).suffix.lower() or ".xlsx"
            inv_path = tmp / f"inventory{inv_ext}"
            out_path = tmp / f"{project}_StackView.xlsx"

            with open(txn_path, "wb") as f:
                f.write(txn_file.getbuffer())
            with open(inv_path, "wb") as f:
                f.write(inv_file.getbuffer())

            brochure_path = None
            if brochure_file:
                brochure_path = tmp / f"brochure{Path(brochure_file.name).suffix.lower() or '.pdf'}"
                with open(brochure_path, "wb") as f:
                    f.write(brochure_file.getbuffer())

            agreement_paths = []
            for i, af in enumerate(agreement_files or []):
                # Keep the real extension -- it decides whether the file is
                # handled as a PDF or as a photo.
                ext = Path(af.name).suffix.lower() or ".pdf"
                ap = tmp / f"agreement_{i}{ext}"
                with open(ap, "wb") as f:
                    f.write(af.getbuffer())
                agreement_paths.append((str(ap), af.name))

            project_dir = PROJECTS_DIR / project

            try:
                result = generate_stack_view(
                    project_dir=str(project_dir),
                    transaction_csv_path=str(txn_path),
                    inventory_path=str(inv_path),
                    output_path=str(out_path),
                    api_key=api_key or None,
                    use_ai_fallback=use_ai,
                    use_ai_inventory=use_ai_inventory,
                    brochure_path=str(brochure_path) if brochure_path else None,
                    agreement_files=agreement_paths,
                    use_ai_documents=use_ai_documents,
                    model=model,
                    tolerance_ft=tolerance_ft,
                )
            except Exception as e:
                st.error(f"Something went wrong: {e}")
                st.stop()

            # Read the output into memory before the temp dir is cleaned up.
            output_bytes = out_path.read_bytes()

    st.success("Stack View generated.")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("No CRE transaction", result.no_txn_count)
    m2.metric("Area missing from text", result.area_missing_count)
    m3.metric("Areas read by AI", result.rows_used_ai)
    m4.metric("Unit nos decoded by AI", result.units_decoded_by_ai)

    if brochure_file or agreement_files:
        e1, e2, e3 = st.columns(3)
        e1.metric("Brochure areas found", result.brochure_areas_found)
        e2.metric("Agreements read", result.agreements_read)
        e3.metric("Agreement vs CRE conflicts", result.agreement_conflicts)
        if result.agreement_conflicts:
            st.info(
                "Conflicting cells are highlighted in Section 1 with the agreement's "
                "figure noted alongside. Section 1 keeps the CRE number by design; "
                "Final Output uses the agreement's."
            )

    if result.per_tower:
        st.write("**Per tower:**")
        st.table({
            "Tower": list(result.per_tower.keys()),
            "No CRE txn": [v["no_txn"] for v in result.per_tower.values()],
            "Area missing": [v["area_missing"] for v in result.per_tower.values()],
        })

    if result.agreement_requests:
        st.info(
            f"**{result.agreement_requests} agreement(s) suggested for download** — one per "
            f"unique area per series, in the 'Agreements to Download' sheet below. "
            f"{len(result.issues)} issue(s) logged in the Review Tracker sheet."
        )
    elif result.issues:
        st.info(f"{len(result.issues)} issue(s) logged in the Review Tracker sheet.")

    if result.warnings:
        with st.expander(f"Processing warnings ({len(result.warnings)})"):
            for w in result.warnings:
                st.warning(w)

    if result.unmatched_transactions:
        with st.expander(f"⚠️ {len(result.unmatched_transactions)} transaction unit numbers couldn't be decoded"):
            st.write(result.unmatched_transactions[:50])

    if result.unmatched_inventory:
        with st.expander(f"⚠️ {len(result.unmatched_inventory)} inventory flat numbers couldn't be decoded"):
            st.write(result.unmatched_inventory[:50])

    st.download_button(
        "⬇️ Download Stack View (.xlsx)",
        data=output_bytes,
        file_name=f"{project}_StackView.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ---- on-screen view of every sheet ----
    st.divider()
    st.subheader("Preview")
    try:
        import io

        import pandas as pd
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(output_bytes), data_only=True)
        tabs = st.tabs(wb.sheetnames)
        for tab, name in zip(tabs, wb.sheetnames):
            with tab:
                ws = wb[name]
                grid = [[("" if c.value is None else c.value) for c in row]
                        for row in ws.iter_rows()]
                if not grid:
                    st.caption("Empty sheet.")
                    continue

                # The list-style sheets have a real header row; the stack view
                # is a laid-out page and is shown as-is. Detect by CONTENT, not
                # by the first cell's text -- the stack view's own details block
                # begins with "Society", which made it look like a table header.
                header_row = None
                for i, row in enumerate(grid[:5]):
                    filled = sum(1 for c in row if str(c).strip())
                    first = str(row[0]).strip().lower() if row else ""
                    if filled >= 8 and first in ("society", "#"):
                        header_row = i
                        break

                if header_row is not None:
                    cols = [str(c) if str(c).strip() else f"col{j}"
                            for j, c in enumerate(grid[header_row])]
                    body = [r for r in grid[header_row + 1:] if any(str(c).strip() for c in r)]
                    df = pd.DataFrame(body, columns=cols).loc[:, [bool(str(c).strip()) for c in grid[header_row]]]
                    st.caption(f"{len(df)} row(s)")
                    st.dataframe(df, use_container_width=True, hide_index=True)
                else:
                    width = max(len(r) for r in grid)
                    padded = [list(r) + [""] * (width - len(r)) for r in grid]
                    st.dataframe(
                        pd.DataFrame(padded, columns=[f"{chr(65 + i) if i < 26 else i}" for i in range(width)]),
                        use_container_width=True, hide_index=True, height=520,
                    )
    except Exception as e:
        st.caption(f"Preview unavailable ({type(e).__name__}: {e}). The download above is unaffected.")
