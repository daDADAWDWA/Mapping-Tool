"""
The two reviewer-facing sheets.

MAPPING REVIEW TRACKER
    Every issue and ambiguity the run found, one row per issue, in the
    reviewer's own tracker layout. This is where the notes live now -- the
    stack grids keep only values and their legend colours, so the grid reads
    as data and the commentary sits somewhere a reviewer can work through,
    sort by severity, and close off.

AGREEMENTS TO BE DOWNLOADED
    The template's checklist requires that "every unique area/type has
    agreement evidence". So for each distinct area in each series that has NO
    agreement behind it, this sheet names ONE flat whose agreement would
    settle it, and says why it matters. One download per unique area, not one
    per flat -- that is the difference between a handful of documents and
    hundreds.

Columns the app cannot know (Raised By, Resolution, review dates, Mapper) are
left empty rather than filled with a guess.
"""

from datetime import date

from openpyxl.styles import Font

REVIEW_TITLE = "MAPPING REVIEW TRACKER"
REVIEW_SUBTITLE = (
    "Reviewer records every issue/comment here. One row per issue. Update "
    "Status and Resolution until closed."
)
REVIEW_COLUMNS = [
    "#", "Date Logged", "Society", "Tower", "Series / Unit", "Floor",
    "Issue Type", "Issue / Comment (Reviewer)", "Raised By", "Severity",
    "Status", "Resolution / Action Taken", "Date Sent for Review", "Reviewed Date",
]

REQUEST_TITLE = "AGREEMENTS TO BE DOWNLOADED"
REQUEST_SUBTITLE = (
    "One agreement per unique area/type per series -- enough to confirm the "
    "area basis for every distinct unit in the tower. Download in priority "
    "order; each row says what it would settle."
)
REQUEST_COLUMNS = [
    "#", "Society", "Tower", "Series", "Unit to download", "Floor",
    "Covers floors", "Current area", "Current area type", "Evidence so far",
    "Why it's needed", "Priority", "Downloaded?", "Notes",
]

# Severity by issue type. High means "this number may be wrong"; Medium means
# "this number is unconfirmed"; Low is housekeeping.
SEVERITY = {
    "Agreement vs CRE": "High",
    "Stated total mismatch": "High",
    "CRE vs RERA": "High",
    "Unverified AI reading": "High",
    "Area differs within series": "Medium",
    "Area not in Marathi text": "Medium",
    "No CRE transaction": "Medium",
    "Area type unconfirmed": "Medium",
    "Brochure only": "Medium",
    "Undecodable unit number": "Low",
    "No data": "Low",
}


def _write(ws, row, values, columns, font):
    for col, name in enumerate(columns, start=1):
        value = values.get(name)
        if value is not None and value != "":
            ws.cell(row=row, column=col, value=value).font = font


def build_review_tracker_sheet(wb, issues, society, sheet_title="Review Tracker"):
    """issues: list of dicts with tower/series/unit/floor/type/comment keys."""
    ws = wb.create_sheet(title=sheet_title)
    bold, normal = Font(name="Arial", bold=True), Font(name="Arial")
    ws.cell(row=1, column=1, value=REVIEW_TITLE).font = bold
    ws.cell(row=2, column=1, value=REVIEW_SUBTITLE).font = normal
    for col, name in enumerate(REVIEW_COLUMNS, start=1):
        ws.cell(row=3, column=col, value=name).font = bold

    today = date.today().strftime("%d/%m/%Y")
    order = {"High": 0, "Medium": 1, "Low": 2}
    ordered = sorted(
        issues,
        key=lambda i: (order.get(SEVERITY.get(i.get("type"), "Low"), 3),
                       str(i.get("series") or ""), i.get("floor") or 0),
    )

    row_num = 4
    for n, issue in enumerate(ordered, start=1):
        _write(ws, row_num, {
            "#": n,
            "Date Logged": today,
            "Society": society,
            "Tower": issue.get("tower") or None,
            "Series / Unit": issue.get("unit") or issue.get("series"),
            "Floor": issue.get("floor"),
            "Issue Type": issue.get("type"),
            "Issue / Comment (Reviewer)": issue.get("comment"),
            "Severity": SEVERITY.get(issue.get("type"), "Medium"),
            "Status": "Open",
        }, REVIEW_COLUMNS, normal)
        row_num += 1

    widths = {"Issue / Comment (Reviewer)": 60, "Resolution / Action Taken": 30,
              "Issue Type": 24, "Society": 20, "Series / Unit": 14}
    for col, name in enumerate(REVIEW_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = \
            widths.get(name, 14)
    ws.freeze_panes = "A4"
    return ws


def build_agreement_requests_sheet(wb, requests, society, sheet_title="Agreements to Download"):
    ws = wb.create_sheet(title=sheet_title)
    bold, normal = Font(name="Arial", bold=True), Font(name="Arial")
    ws.cell(row=1, column=1, value=REQUEST_TITLE).font = bold
    ws.cell(row=2, column=1, value=REQUEST_SUBTITLE).font = normal
    for col, name in enumerate(REQUEST_COLUMNS, start=1):
        ws.cell(row=3, column=col, value=name).font = bold

    order = {"High": 0, "Medium": 1, "Low": 2}
    ordered = sorted(requests, key=lambda r: (order.get(r.get("priority"), 3),
                                              str(r.get("series") or "")))
    row_num = 4
    for n, req in enumerate(ordered, start=1):
        _write(ws, row_num, {
            "#": n,
            "Society": society,
            "Tower": req.get("tower") or None,
            "Series": req.get("series"),
            "Unit to download": req.get("unit"),
            "Floor": req.get("floor"),
            "Covers floors": req.get("floors"),
            "Current area": req.get("area"),
            "Current area type": req.get("area_type"),
            "Evidence so far": req.get("evidence") or "none",
            "Why it's needed": req.get("reason"),
            "Priority": req.get("priority"),
            "Downloaded?": "No",
            "Notes": req.get("notes"),
        }, REQUEST_COLUMNS, normal)
        row_num += 1

    widths = {"Why it's needed": 60, "Notes": 34, "Current area type": 18,
              "Society": 20, "Covers floors": 18}
    for col, name in enumerate(REQUEST_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = \
            widths.get(name, 14)
    ws.freeze_panes = "A4"
    return ws


def suggest_agreements(final_rows, tower_entries, tolerance_ft=5.0):
    """
    Which agreements to ask for. One per distinct area per series that has no
    agreement behind it.

    The flat suggested is a floor in that row which HAS a registered
    transaction -- an agreement can only be downloaded if a registration
    exists. Where no floor in the row has one, that is said plainly instead of
    naming a flat whose agreement can't be found.
    """
    # floors that have a CRE transaction, per tower+series column
    cre_floors = {}
    for entry in tower_entries:
        tm = entry["tm"]
        for position, col in enumerate(tm.section1.series_cols, start=1):
            floors = sorted(
                e["floor"] for e in entry["by_col_values"].get(col, [])
                if e.get("value") is not None
            )
            cre_floors[(entry["letter"], position)] = floors

    requests = []
    for row in final_rows:
        if row.get("agreement_backed"):
            continue
        floors = row.get("floor_list") or []
        if not floors:
            continue
        position = row.get("position")
        available = [f for f in cre_floors.get((row.get("tower_letter"), position), [])
                     if f in floors]

        if available:
            # Prefer a middle floor: the first and last floors of a stack are
            # the ones most likely to be atypical (podium setbacks, terraces),
            # so a mid-stack flat is the better evidence for the whole group.
            unit_floor = available[len(available) // 2]
            unit = f"{unit_floor * 100 + position}"
            notes = None
        else:
            unit_floor = floors[len(floors) // 2]
            unit = f"{unit_floor * 100 + position}"
            notes = ("No registered transaction on these floors, so an agreement "
                     "may not be obtainable — check with the builder.")

        n_floors = len(floors)
        if n_floors >= 5:
            priority = "High"
        elif n_floors >= 2:
            priority = "Medium"
        else:
            priority = "Low"

        requests.append({
            "tower": row.get("tower") or None,
            "series": row.get("series"),
            "unit": unit,
            "floor": unit_floor,
            "floors": row.get("floors"),
            "area": row.get("area"),
            "area_type": row.get("area_type"),
            "evidence": row.get("evidence"),
            "priority": priority,
            "notes": notes,
            "reason": (
                f"{n_floors} flat(s) share this area and none has agreement evidence, "
                f"so the area basis ({row.get('area_type') or 'unknown'}) is unconfirmed."
            ),
        })
    return requests


def draft_reasons_with_ai(requests, api_key=None, model=None):
    """
    Rewrite the "Why it's needed" lines as one short sentence each, in a single
    batched call. Purely cosmetic: if it fails, the deterministic wording
    stays, so nothing depends on it.
    """
    if not requests:
        return []
    from .ai_assist import _ask, _client
    client = _client(api_key)
    if client is None:
        return []

    listing = "\n".join(
        f"{i}. Series {r['series']}, floors {r['floors']}, area {r['area']}, "
        f"type {r['area_type'] or 'unknown'}, evidence {r['evidence'] or 'none'}"
        for i, r in enumerate(requests, start=1)
    )
    prompt = (
        "A property-mapping reviewer needs to download flat agreements to "
        "confirm carpet areas. For each item below, write ONE short sentence "
        "(under 25 words) saying what downloading that agreement would settle. "
        "Be concrete and avoid repeating the numbers back.\n\n"
        f"{listing}\n\n"
        'Return ONLY JSON: {"reasons": ["...", "..."]} with one entry per item, '
        "in the same order. No other text."
    )
    data = _ask(client, prompt, max_tokens=1500, model=model)
    if not isinstance(data, dict):
        return []
    reasons = data.get("reasons")
    if not isinstance(reasons, list) or len(reasons) != len(requests):
        return []
    return [str(r).strip() for r in reasons]
