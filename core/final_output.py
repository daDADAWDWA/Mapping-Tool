"""
Builds the "Final Output -- NocoDB Hand-off" sheet: one row per distinct area
within each Series per Tower, with a reconciled, evidence-backed value.

SOURCE PRIORITY (highest first)
==============================
    1. AGREEMENT        the registered agreement for that specific flat
    2. CRE TRANSACTION  the registration record
    3. RERA INVENTORY   the builder's MahaRERA disclosure
    4. BROCHURE         the developer's marketing brochure

Resolution is PER FLOOR, not per series: each floor takes its value from the
highest-priority source that has one for that floor. An agreement applies to
its OWN FLAT ONLY -- it does not propagate across the series. So an agreement
whose area matches its neighbours within tolerance simply merges into their
row, and only one that genuinely differs splits off a row of its own.

Floors are then grouped into rows by area, with near-duplicate values (within
`tolerance_ft`, default 5 sq ft) treated as the same area -- measurement
noise, not a real difference. The reading appearing on the most floors wins;
ties go to the larger value. Floor ranges can be non-contiguous ("4-6,8-15").

Lower-priority sources are never silently dropped: wherever one disagrees
with the winning value beyond tolerance, Mapping Status names the source and
the exact floors, so a reviewer can see what was overruled. Area type comes
from the winning source, and a change of measurement BASIS (RERA carpet vs
MOFA carpet vs built-up) is flagged separately, since that matters more than
a couple of square feet.

Final Carpet (ft²) is always a straight conversion of that row's m² figure --
never an independently reported number.
"""

import re
from collections import Counter, defaultdict

from openpyxl.styles import Font

FT_PER_M2 = 10.7639104167
DEFAULT_TOLERANCE_FT = 5.0

HEADER_TITLE = "FINAL OUTPUT — NocoDB HAND-OFF"
HEADER_SUBTITLE = (
    "Finalized, evidence-backed values per series/unit-type. These map to the "
    "NocoDB fields: series, areas, floor offset, exit direction, OC and map "
    "link. Fill only after the Reviewer signs off."
)
COLUMNS = [
    "Society", "Tower", "Series", "Unit Type", "Applicable Floors",
    "Final Carpet (m²)", "Final Carpet (ft²)", "Area Type", "Balcony Area",
    "Floor Offset", "Exit Direction", "OC Status", "Map Link", "Evidence Ref",
    "Mapping Status", "Finalized By", "Finalized Date",
]

# When a project has no tower/wing split, the verified hand-off writes this
# rather than leaving Tower blank.
STANDALONE_TOWER = "Standalone"


def format_series_label(raw):
    """'Series 1' -> '01', matching the verified hand-off sheet."""
    digits = re.sub(r'\D', '', str(raw or ""))
    return digits.zfill(2) if digits else (str(raw).strip() or None)


def format_m2(value):
    """115.21 -> '115.21 m²'; 91.5 -> '91.5 m²' (no trailing zeros)."""
    if value is None:
        return None
    s = f"{value:.2f}".rstrip("0").rstrip(".")
    return f"{s} m²"


def format_ft2(value_m2):
    """m² -> ' 1,240 ft²' -- a straight conversion, rounded to whole feet."""
    if value_m2 is None:
        return None
    return f"{value_m2 * FT_PER_M2:,.0f} ft²"


# Priority order, best first. Drives both resolution and reporting.
PRIORITY = ["Agreement", "CRE", "RERA", "Brochure"]

AREA_TYPE_DISPLAY = {
    "reracarpet": "RERA Carpet",
    "mofacarpet": "MOFA Carpet",
    "builtup": "Built-up",
    "superbuiltup": "Super Built-up",
    "saleable": "Saleable",
    "carpet": "Carpet",
}

# Free-text area wording -> compact label. Most specific first.
_AREA_TYPE_PHRASES = [
    ("super built", "superbuiltup"), ("superbuilt", "superbuiltup"),
    ("सुपर बिल्ट", "superbuiltup"),
    ("mofa", "mofacarpet"), ("मोफा", "mofacarpet"),
    ("rera carpet", "reracarpet"), ("रेरा कार्पेट", "reracarpet"),
    ("built up", "builtup"), ("built-up", "builtup"), ("builtup", "builtup"),
    ("बिल्ट अप", "builtup"),
    ("saleable", "saleable"), ("विक्रीयोग्य", "saleable"),
    ("carpet", "carpet"), ("कार्पेट", "carpet"),
]

_SQM_UNIT_FORMS = {"sqm", "sqmt", "sqmtr", "sqmtrs", "sqmeter", "sqmeters",
                   "sqmetre", "sqmetres", "m2", "m²"}


def to_m2(value, unit):
    """Any stated area -> m². None if the value isn't numeric."""
    if value is None:
        return None
    try:
        value = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    unit_norm = re.sub(r'[^a-z0-9²]', '', str(unit or "").lower())
    return value if unit_norm in _SQM_UNIT_FORMS else value / FT_PER_M2


def _cre_value_in_m2(entry):
    """A CRE grid entry -> m². CRE descriptions state sq.ft or sq.m."""
    return to_m2(entry.get("value"), entry.get("unit"))


def normalise_area_type(text):
    """Free-text area wording -> a display label, or None."""
    if not text:
        return None
    low = str(text).lower()
    for phrase, label in _AREA_TYPE_PHRASES:
        if phrase in low or phrase in str(text):
            return AREA_TYPE_DISPLAY[label]
    return str(text).strip().title() or None


def format_floor_ranges(floors):
    """[4,5,6,8,...,15] -> '4-6,8-15'; [1] -> '1'; [] -> None."""
    if not floors:
        return None
    floors = sorted(set(floors))
    segments, start, prev = [], floors[0], floors[0]
    for f in floors[1:]:
        if f == prev + 1:
            prev = f
            continue
        segments.append((start, prev))
        start = prev = f
    segments.append((start, prev))
    return ",".join(str(a) if a == b else f"{a}-{b}" for a, b in segments)


def parse_floor_ranges(text):
    """'3-7,9-14 ,16' -> [3,4,5,6,7,9,...,14,16]. [] if unparseable."""
    if not text:
        return []
    floors = []
    for part in str(text).replace(" ", "").split(","):
        if not part:
            continue
        m = re.fullmatch(r'(\d+)-(\d+)', part)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            floors.extend(range(min(a, b), max(a, b) + 1))
        elif part.isdigit():
            floors.append(int(part))
    return sorted(set(floors))


def cluster_with_floors(floor_value_pairs, tolerance, priority_of_floor=None):
    """
    Groups near-duplicate VALUES (within `tolerance`) into clusters. One entry
    per cluster: {"floors": [...], "representative": value}, where the
    representative is the exact value on the most floors (ties -> larger).
    """
    pairs = [(f, v) for f, v in floor_value_pairs if v is not None]
    if not pairs:
        return []

    floors_by_value = defaultdict(list)
    for floor, value in pairs:
        floors_by_value[value].append(floor)

    distinct_sorted = sorted(floors_by_value)
    value_clusters, current = [], [distinct_sorted[0]]
    for v in distinct_sorted[1:]:
        if v - current[-1] < tolerance:
            current.append(v)
        else:
            value_clusters.append(current)
            current = [v]
    value_clusters.append(current)

    results = []
    for cluster_values in value_clusters:
        all_floors = []
        for v in cluster_values:
            all_floors.extend(floors_by_value[v])

        # Which exact reading represents the cluster? SOURCE PRIORITY decides
        # first. Choosing purely by "appears on the most floors" silently
        # defeated the whole priority chain: a RERA value on 15 floors beat an
        # agreement/CRE value on 7, so Section 2 overrode Section 1 even though
        # Section 1 outranks it.
        def rank(v):
            """
            How well is this exact reading supported, by source quality?

            Counts the floors backing it AT EACH PRIORITY LEVEL and compares
            those counts in priority order. So a value with 1 agreement floor
            and 6 CRE floors beats one with 1 agreement floor and 7 RERA
            floors -- the second has more floors overall, but its support comes
            from a weaker source. Comparing raw floor counts let a RERA value
            override a CRE one; comparing per-level counts is what actually
            respects the priority chain.
            """
            if not priority_of_floor:
                return (0, -len(floors_by_value[v]), -v)
            counts = [0] * (len(PRIORITY) + 1)
            for f in floors_by_value[v]:
                counts[min(priority_of_floor.get(f, len(PRIORITY)), len(PRIORITY))] += 1
            return tuple(-c for c in counts) + (-v,)

        representative = sorted(cluster_values, key=rank)[0]
        results.append({"floors": sorted(set(all_floors)), "representative": representative})
    return results


def _get_society_name(tm_rows, fallback_name):
    for row in tm_rows:
        if row and row[0] and row[0].strip().lower() == "society":
            # Templates pack two "Label , , , Value" pairs per row, so only
            # scan the left half -- the whole row risks picking up the second
            # pair's LABEL (e.g. "Mapper") as if it were this field's value.
            midpoint = max(1, len(row) // 2)
            for cell in row[1:midpoint]:
                if cell and cell.strip():
                    return cell.strip()
            return fallback_name
    return fallback_name


def _majority(values):
    values = [v for v in values if v]
    return Counter(values).most_common(1)[0][0] if values else None


def build_per_floor_sources(floors, cre_entries, rera_entries,
                            agreements_by_pos, position, brochure_entry):
    """
    For one series column: {floor: {source_name: {value_m, area_type,
    balcony, evidence, verified}}} covering every source with a value.
    """
    by_floor = defaultdict(dict)

    for e in rera_entries:
        if e.get("value") is not None and e["floor"] in floors:
            by_floor[e["floor"]]["RERA"] = {
                "value_m": e["value"], "area_type": "RERA Carpet",
                "balcony": None, "evidence": None,
                "verified": e.get("verified", True),
            }

    for e in cre_entries:
        v = _cre_value_in_m2(e)
        if v is not None and e["floor"] in floors:
            compact = e.get("area_type")
            # A type confirmed by matching an agreement's area beats one merely
            # parsed out of the Marathi wording.
            confirmed = e.get("confirmed_area_type")
            by_floor[e["floor"]]["CRE"] = {
                "value_m": v,
                "area_type": confirmed or AREA_TYPE_DISPLAY.get(
                    compact, compact.title() if compact else None),
                # The CRE value_m is carpet PLUS balcony where the description
                # stated one; the balcony is also reported on its own here.
                "balcony": e.get("balcony"), "evidence": None, "verified": True,
            }

    for (floor, pos), rec in agreements_by_pos.items():
        if pos != position or floor not in floors:
            continue
        # A balcony-only agreement contributes NO carpet area -- only its
        # balcony figure. Letting it through as a carpet value would put a
        # ~80 sq ft balcony where a 1200 sq ft flat should be.
        # An agreement states RERA carpet EXCLUDING balcony. CRE figures already
        # have their balcony folded in upstream, so an agreement must get the
        # same treatment or it reads ~10 m² short and, being the top-priority
        # source, drags the whole row down with it.
        # An agreement reader that has already worked out the flat's area --
        # deciding whether the balcony was inside the figure or needed adding --
        # hands it over resolved. Re-adding the balcony here would double-count
        # a deed that states its carpet figure as a total already.
        balcony = rec.get("balcony_area")
        if rec.get("final_area_m2") is not None:
            v = rec["final_area_m2"]
        else:
            v = to_m2(rec.get("area_value"), rec.get("area_unit"))
            if v is not None and balcony is not None:
                balcony_m = to_m2(balcony, rec.get("area_unit"))
                if balcony_m is not None:
                    v += balcony_m
        if v is None:
            if balcony is not None:
                by_floor[floor]["AgreementBalcony"] = {
                    "value_m": None, "area_type": None, "balcony": balcony,
                    "evidence": rec.get("filename"),
                    "verified": rec.get("verified", True),
                }
            continue
        by_floor[floor]["Agreement"] = {
            "combined": rec.get("combined"),
            "value_m": v,
            "area_type": normalise_area_type(rec.get("area_type")),
            "balcony": balcony,
            "evidence": (f"{rec.get('filename')} p.{rec['page']}"
                         if rec.get("page") else rec.get("filename")),
            "verified": rec.get("verified", True),
        }

    # Brochure last, because its scope depends on what the other sources know.
    # A brochure that doesn't state which floors its plan covers is NOT spread
    # across every floor in the tower -- that would invent a range the
    # document never claimed. It is limited to floors some other source has
    # already shown to be real units; if there are none, it is reported as an
    # unscoped brochure-only figure instead.
    brochure_unscoped = False
    if brochure_entry:
        b_m = to_m2(brochure_entry.get("carpet_area"), brochure_entry.get("carpet_unit"))
        if b_m is not None:
            stated = [f for f in parse_floor_ranges(brochure_entry.get("applicable_floors"))
                      if f in floors]
            targets = stated or sorted(by_floor)
            brochure_unscoped = not stated and not targets
            for f in targets:
                by_floor[f]["Brochure"] = {
                    "value_m": b_m,
                    "area_type": normalise_area_type(brochure_entry.get("area_type")),
                    "balcony": brochure_entry.get("balcony_area"),
                    "evidence": f"Brochure p.{brochure_entry.get('page')}",
                    "verified": brochure_entry.get("verified", True),
                }

    return by_floor, brochure_unscoped


def build_final_output_sheet(wb, tower_entries, fallback_society_name,
                             tolerance_ft=DEFAULT_TOLERANCE_FT):
    """
    tower_entries: one dict per tower, each with:
        {"letter", "tm", "by_col_values", "rera_by_col_values",
         "agreements_by_pos", "brochure_by_col"}
    """
    tolerance_m = tolerance_ft / FT_PER_M2
    ws = wb.create_sheet(title="Final Output")
    bold = Font(name="Arial", bold=True)
    normal = Font(name="Arial")

    ws.cell(row=1, column=1, value=HEADER_TITLE).font = bold
    ws.cell(row=2, column=1, value=HEADER_SUBTITLE).font = normal
    for c, name in enumerate(COLUMNS, start=1):
        ws.cell(row=3, column=c, value=name).font = bold

    emitted_rows = []           # returned for the review + agreement sheets
    row_num = 4
    for entry in tower_entries:
        tm = entry["tm"]
        letter = entry["letter"]
        society = _get_society_name(tm.rows, fallback_society_name)
        agreements_by_pos = entry.get("agreements_by_pos", {})
        brochure_by_col = entry.get("brochure_by_col", {})
        floors = set(tm.section1.floor_rows) | set(tm.section2.floor_rows)

        for position, col_idx in enumerate(tm.section1.series_cols, start=1):
            series_label = format_series_label(tm.rows[tm.section1.header_row][col_idx])
            # Unit Type comes from the brochure when there is one. The
            # template's "Unit Type ->" row holds illustrative labels
            # ("4 BHK Signature"), not this project's real mix, so copying it
            # would state something unverified as fact.
            brochure_entry = brochure_by_col.get(col_idx)
            unit_type = (brochure_entry or {}).get("unit_type")

            by_floor, brochure_unscoped = build_per_floor_sources(
                floors,
                entry["by_col_values"].get(col_idx, []),
                entry["rera_by_col_values"].get(col_idx, []),
                agreements_by_pos, position,
                brochure_entry,
            )

            # Winner per floor = highest-priority source present.
            winners = {}
            for floor, sources in by_floor.items():
                for name in PRIORITY:          # note: excludes AgreementBalcony
                    if name in sources:
                        winners[floor] = (name, sources[name])
                        break

            if not winners:
                # No source has anything for this series -- the template simply
                # has more series columns than the project has flats per floor.
                # A blank row here is noise in the hand-off, so skip it.
                if not (brochure_unscoped and brochure_entry):
                    continue
                row = {
                    "Society": society, "Tower": letter or STANDALONE_TOWER,
                    "Series": series_label, "Unit Type": unit_type or None,
                    "Mapping Status": "In Progress",
                }
                if brochure_unscoped and brochure_entry:
                    b_m = to_m2(brochure_entry.get("carpet_area"),
                                brochure_entry.get("carpet_unit"))
                    if b_m is not None:
                        row.update({
                            "Final Carpet (m²)": format_m2(b_m),
                            "Final Carpet (ft²)": format_ft2(b_m),
                            "Area Type": normalise_area_type(brochure_entry.get("area_type")),
                            "Evidence Ref": f"Brochure p.{brochure_entry.get('page')}",
                            "Mapping Status": "Needs Review — brochure only; no floors stated "
                                              "and no other source covers this series",
                        })
                _write_row(ws, row_num, normal, row)
                row_num += 1
                continue

            # A DUPLEX is one home over two floors. It becomes a single row --
            # floors joined with "/" and the two floors' areas summed -- and its
            # floors are removed from the ordinary clustering so they don't also
            # appear in the series' normal row.
            duplex_rows, spans = [], {}
            for f, (src, info) in winners.items():
                combined = info.get("combined") if isinstance(info, dict) else None
                if src == "Agreement" and combined and combined.get("kind") == "duplex":
                    spans.setdefault(tuple(combined["floors"]), set()).add(f)
            for span in spans:
                present = [f for f in span if f in winners]
                if len(present) < 2:
                    continue
                total_m = sum(winners[f][1]["value_m"] for f in present)
                info = winners[present[0]][1]
                duplex_rows.append({
                    "first_floor": min(present),
                    "row": {
                        "Society": society,
                        "Tower": letter or STANDALONE_TOWER,
                        "Series": series_label,
                        "Unit Type": "Duplex",
                        "Applicable Floors": "/".join(str(f) for f in sorted(present)),
                        "Final Carpet (m²)": format_m2(total_m),
                        "Final Carpet (ft²)": format_ft2(total_m),
                        "Area Type": info.get("area_type"),
                        "Evidence Ref": info.get("evidence"),
                        "Mapping Status": "In Progress",
                    },
                    "meta": {
                        "tower": letter or STANDALONE_TOWER, "tower_letter": letter,
                        "position": position, "series": series_label,
                        "floors": "/".join(str(f) for f in sorted(present)),
                        "floor_list": list(present), "area": format_m2(total_m),
                        "area_type": info.get("area_type"),
                        "evidence": info.get("evidence"), "status": "In Progress",
                        "agreement_backed": True,
                    },
                })
                for f in present:
                    winners.pop(f, None)

            if not winners and not duplex_rows:
                continue

            priority_of_floor = {f: PRIORITY.index(w[0]) for f, w in winners.items()}
            clusters = cluster_with_floors(
                [(f, w[1]["value_m"]) for f, w in winners.items()], tolerance_m,
                priority_of_floor=priority_of_floor,
            )

            emitted = list(duplex_rows)
            for cluster in clusters:
                cf = cluster["floors"]
                final_m = cluster["representative"]
                cluster_winners = [winners[f] for f in cf]
                win_source = _majority([w[0] for w in cluster_winners]) or "CRE"

                # Anything lower-priority that disagrees beyond tolerance.
                # Sources disagreeing on the SAME floors are reported together,
                # so the common case reads "CRE, RERA, Brochure differ on
                # floor 12" rather than three near-identical clauses.
                disagreements = defaultdict(list)
                for name in PRIORITY[PRIORITY.index(win_source) + 1:]:
                    off = sorted(
                        f for f in cf
                        if name in by_floor[f]
                        and abs(by_floor[f][name]["value_m"] - final_m) >= tolerance_m
                    )
                    if off:
                        disagreements[format_floor_ranges(off)].append(name)
                notes = [
                    f"{', '.join(names)} differ{'s' if len(names) == 1 else ''} "
                    f"on floor(s) {rng}"
                    for rng, names in disagreements.items()
                ]

                area_type = _majority([w[1]["area_type"] for w in cluster_winners])
                other_types = {
                    by_floor[f][n]["area_type"]
                    for f in cf for n in by_floor[f] if n in PRIORITY
                    if by_floor[f][n]["area_type"] and by_floor[f][n]["area_type"] != area_type
                }
                if other_types:
                    notes.append("area type differs: " + ", ".join(sorted(other_types)))

                if any(w[1].get("verified") is False for w in cluster_winners):
                    notes.append("area read from an image, unverified")

                status = f"Needs Review — {'; '.join(notes)}" if notes else "In Progress"
                # Balcony is a separate measurement, so it comes from the best
                # source that actually states one -- not from whichever source
                # happened to win the carpet area.
                balcony = None
                for name in ["AgreementBalcony"] + PRIORITY:
                    stated = [by_floor[f][name].get("balcony") for f in cf
                              if name in by_floor[f]
                              and by_floor[f][name].get("balcony") is not None]
                    if stated:
                        balcony = _majority(stated)
                        break
                evidence = "; ".join(sorted({
                    w[1]["evidence"] for w in cluster_winners if w[1].get("evidence")
                })) or None

                emitted.append({
                    "first_floor": cf[0],
                    "meta": {
                        "tower": letter or STANDALONE_TOWER,
                        "tower_letter": letter, "position": position,
                        "series": series_label, "floors": format_floor_ranges(cf),
                        "floor_list": list(cf), "area": format_m2(final_m),
                        "area_type": area_type, "evidence": evidence,
                        "status": status,
                        "agreement_backed": any(w[0] == "Agreement" for w in cluster_winners),
                    },
                    "row": {
                        "Society": society,
                        "Tower": letter or STANDALONE_TOWER,
                        "Series": series_label,
                        "Unit Type": unit_type or None,
                        "Applicable Floors": format_floor_ranges(cf),
                        "Final Carpet (m²)": format_m2(final_m),
                        "Final Carpet (ft²)": format_ft2(final_m),
                        "Area Type": area_type,
                        # Balcony is already inside Final Carpet, so repeating
                        # it here would read as an extra area on top.
                        "Evidence Ref": evidence,
                        "Mapping Status": status,
                    },
                })

            # Rows within a series read top-to-bottom by floor, as the
            # verified sheet does -- clusters come out in value order otherwise.
            for item in sorted(emitted, key=lambda e: e["first_floor"]):
                _write_row(ws, row_num, normal, item["row"])
                if item.get("meta"):
                    emitted_rows.append(item["meta"])
                row_num += 1

    for c in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[ws.cell(row=3, column=c).column_letter].width = 16
    return ws, emitted_rows


def _write_row(ws, row_num, font, values):
    for c, name in enumerate(COLUMNS, start=1):
        val = values.get(name)
        if val is not None:
            ws.cell(row=row_num, column=c, value=val).font = font
