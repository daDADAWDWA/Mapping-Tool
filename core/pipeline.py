"""
Main orchestration: takes a PROJECT (a folder containing either one
template.csv directly, or one subfolder per tower each with its own
template.csv) + a transaction CSV + an inventory file (Excel OR a MahaRERA-
style "Sold/Booked Inventory" disclosure PDF), and produces a filled-in
.xlsx.

Single-tower projects (just projects/<Name>/template.csv) get one sheet,
exactly as before.

Multi-tower projects (projects/<Name>/A/template.csv, /B/template.csv, ...)
get the Tower/Wing column read from both input files, split by tower, and
one sheet generated per tower -- each tower can have a completely different
floor/series layout, since each has its own template file.
"""

import re
from collections import defaultdict, Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from .template_model import TemplateModel
from .decode import decode_unit, normalize_tower
from .area_extract import extract_area
from .cleaning import clean_transaction_df, clean_inventory_df
from .aliases import load_aliases, find_column
from .inventory_pdf import extract_inventory_from_pdf
from .ai_assist import decode_units_ai, resolve_columns_ai
from .final_output import build_final_output_sheet, to_m2, FT_PER_M2
from .doc_extract import extract_brochure, extract_agreements
from .config import load_settings
from .area_type import confirm_area_types, summarise_series_type
from .area_extract import primary_area
from .numbering import (
    FLOOR_POSITION,
    collect_samples,
    decode as decode_with_rule,
    floor_from_text,
    infer_rule,
)
from .txn_normalise import extract_unit_tower
from .review import (
    build_agreement_requests_sheet,
    build_review_tracker_sheet,
    draft_reasons_with_ai,
    suggest_agreements,
)
from . import colors as C


class ProcessingResult:
    def __init__(self):
        self.output_path = None
        self.warnings = []
        self.unmatched_transactions = []
        self.unmatched_inventory = []
        self.no_txn_count = 0
        self.area_missing_count = 0
        self.rows_used_ai = 0
        self.units_decoded_by_ai = 0
        self.brochure_areas_found = 0
        self.agreements_read = 0
        self.agreement_conflicts = 0
        self.issues = []          # every ambiguity, for the Review Tracker
        self.agreement_requests = 0
        self.per_tower = {}  # {tower_label: {"no_txn": n, "area_missing": n, "rows_used_ai": n}}


def build_inventory_grid(inventory_df: pd.DataFrame, aliases: dict, result: ProcessingResult,
                         tower_label="", tm=None, api_key=None, use_ai_decode=False,
                         model=None):
    flat_col = find_column(inventory_df.columns, "flat_no_inventory", aliases)
    area_col = find_column(inventory_df.columns, "carpet_area", aliases)

    flat_col, area_col = _resolve_missing_columns(
        inventory_df, {"flat_no_inventory": flat_col, "carpet_area": area_col},
        result, tower_label, api_key, model, use_ai_decode, "inventory file",
    )

    if flat_col is None or area_col is None:
        prefix = f"[{tower_label}] " if tower_label else ""
        missing = " and ".join(
            n for n, c in (("Flat No", flat_col), ("Carpet Area", area_col)) if c is None
        )
        result.warnings.append(
            f"{prefix}Inventory file: could not find the {missing} column (found columns: "
            f"{[c for c in inventory_df.columns if not str(c).startswith('_')]}). Section 2 "
            f"will be left empty. Turn on the AI switch for column matching, or add a "
            f"column_aliases.json next to the template -- see README."
        )
        return {}, 0, set()

    # Floors this tower actually has. A decoded floor outside this list is a
    # mis-decode, not a real unit -- e.g. the jodi "1601/1602" collapses to
    # 16011602 under the plain rule, which looks like floor 160116. Without
    # this check such rows land in the grid, match no template row, and
    # disappear without ever being reported.
    known_floors = set(tm.section2.floor_rows) if tm is not None else None

    grid = {}
    unverified = set()    # (floor, position) read by AI but never cross-checked
    pending = []          # (raw_identifier, area) the plain rule couldn't decode
    has_verified = "verified" in inventory_df.columns
    for _, row in inventory_df.iterrows():
        raw = row[flat_col]
        try:
            area = float(row[area_col])
        except (ValueError, TypeError):
            continue
        row_verified = _truthy(row["verified"]) if has_verified else True
        decoded = decode_unit(raw)
        if decoded is None or (known_floors is not None and decoded[0] not in known_floors):
            if str(raw).strip():
                pending.append((str(raw).strip(), area))
            continue
        grid[decoded] = area
        if not row_verified:
            unverified.add(decoded)

    decoded_by_ai = 0
    if pending and use_ai_decode and tm is not None:
        section = tm.section2
        resolved = decode_units_ai(
            [raw for raw, _ in pending],
            floors=set(section.floor_rows),
            num_series=section.num_series,
            api_key=api_key,
            series_labels=[tm.rows[section.header_row][c] for c in section.series_cols],
            source="builder inventory",
            model=model,
        )
        for raw, area in pending:
            hit = resolved.get(raw)
            if hit is not None:
                grid[(hit["floor"], hit["position"])] = area
                decoded_by_ai += 1
            else:
                result.unmatched_inventory.append(raw)
    else:
        result.unmatched_inventory.extend(raw for raw, _ in pending)

    return grid, decoded_by_ai, unverified


def build_transaction_grid(transaction_df: pd.DataFrame, aliases: dict, result: ProcessingResult,
                            api_key: str = None, use_ai_fallback: bool = True, tower_label="",
                            tm=None, use_ai_decode: bool = False, model: str = None):
    unit_col = find_column(transaction_df.columns, "unit_no", aliases)
    desc_col = find_column(transaction_df.columns, "description", aliases)

    unit_col, desc_col = _resolve_missing_columns(
        transaction_df, {"unit_no": unit_col, "description": desc_col},
        result, tower_label, api_key, model, use_ai_decode, "transaction file",
    )

    if unit_col is None or desc_col is None:
        prefix = f"[{tower_label}] " if tower_label else ""
        # Name only what's actually missing -- saying "both" when one WAS found
        # sends you hunting for the wrong problem.
        missing = " and ".join(
            n for n, c in (("Unit No", unit_col), ("Property Description", desc_col))
            if c is None
        )
        result.warnings.append(
            f"{prefix}Transaction file: could not find the {missing} column (found columns: "
            f"{[c for c in transaction_df.columns if not str(c).startswith('_')]}). Section 1 "
            f"will be left empty. Turn on the AI switch for column matching, or add a "
            f"column_aliases.json next to the template -- see README."
        )
        return {}, 0, 0

    date_col = "_reg_date_parsed" if "_reg_date_parsed" in transaction_df.columns else None
    year_col = find_column(transaction_df.columns, "registration_year", aliases)

    known_floors = set(tm.section1.floor_rows) if tm is not None else None
    prefix_label = f"[{tower_label}] " if tower_label else ""

    # enrich_transactions() has already read every description and resolved the
    # unit, floor, series and area -- including which BASIS that area is on.
    # Falling back to decode_unit() here would undo the inferred numbering rule.
    enriched = "_floor_resolved" in transaction_df.columns

    records = []
    pending = []          # (raw_identifier, sort_key, year, description)
    for idx, row in transaction_df.iterrows():
        sort_key = row[date_col] if date_col else pd.NaT
        year_val = row[year_col] if year_col else None
        raw_unit = str(row["_unit_resolved"] if enriched else row[unit_col] or "").strip()

        if enriched:
            floor, position = row["_floor_resolved"], row["_series_resolved"]
            decoded = ((int(floor), int(position))
                       if pd.notna(floor) and pd.notna(position) else None)
        else:
            decoded = decode_unit(row[unit_col])

        if decoded is None or (known_floors is not None and decoded[0] not in known_floors):
            if raw_unit:
                pending.append((raw_unit, sort_key, year_val, row[desc_col]))
            continue
        records.append((decoded, sort_key, year_val, row[desc_col]))

    decoded_by_ai = 0
    if pending and use_ai_decode and tm is not None:
        section = tm.section1
        resolved = decode_units_ai(
            [raw for raw, _, _, _ in pending],
            floors=set(section.floor_rows),
            num_series=section.num_series,
            api_key=api_key,
            series_labels=[tm.rows[section.header_row][c] for c in section.series_cols],
            source="CRE transaction",
            model=model,
        )
        for raw, sort_key, year_val, desc in pending:
            hit = resolved.get(raw)
            if hit is not None:
                records.append(((hit["floor"], hit["position"]), sort_key, year_val, desc))
                decoded_by_ai += 1
            else:
                result.unmatched_transactions.append(raw)
    else:
        result.unmatched_transactions.extend(raw for raw, _, _, _ in pending)

    best_by_cell = {}
    for decoded, sort_key, year_val, desc in records:
        prev = best_by_cell.get(decoded)
        is_better = False
        if prev is None:
            is_better = True
        else:
            prev_sort_key, prev_year, _ = prev
            if pd.notna(sort_key) and pd.notna(prev_sort_key):
                is_better = sort_key > prev_sort_key
            elif pd.notna(sort_key) and pd.isna(prev_sort_key):
                is_better = True
            elif pd.isna(sort_key) and pd.isna(prev_sort_key):
                try:
                    is_better = float(year_val) > float(prev_year)
                except (TypeError, ValueError):
                    is_better = False
        if is_better:
            best_by_cell[decoded] = (sort_key, year_val, desc)

    grid = {}
    rows_used_ai = 0
    for decoded, (sort_key, year_val, desc) in best_by_cell.items():
        area = extract_area(
            desc, api_key=api_key, use_ai_fallback=use_ai_fallback, model=model
        )
        value, unit, area_type = area["value"], area["unit"], area["area_type"]

        # Any stated basis is acceptable -- carpet where it exists, otherwise
        # built-up, super built-up or saleable -- as long as the TYPE travels
        # with the number. Refusing anything but carpet left 43 of 47 Embassy
        # rows empty, because those deeds are written on super built-up.
        if value is None:
            any_value, any_unit, any_label = primary_area(desc)
            if any_value is not None:
                value, unit, area_type = any_value, any_unit, any_label
                area = dict(area, value=any_value, unit=any_unit,
                            area_type=any_label, source="regex")
        balcony, balcony_unit = area["balcony"], area["balcony_unit"]
        if area["source"] == "ai":
            rows_used_ai += 1

        # RERA carpet EXCLUDES the balcony, so the flat's real area is carpet
        # plus balcony -- UNLESS the description already writes out the total
        # itself, in which case adding the balcony again double-counts it.
        # A real description reads:
        #   "क्षेत्र 1562 चौ फुट रेरा कार्पेट व बाल्कनी क्षेत्र 128 चौ फुट
        #    अशाप्रकारे ... एकूण क्षेत्र 1690 चौ फुट"
        # 1562 + 128 = 1690 is already stated; summing again gave 1818.
        carpet_only = value
        stated_total = area.get("stated_total")
        total_unit = area.get("total_unit") or unit
        if stated_total is not None and value is not None:
            total_m2 = to_m2(stated_total, total_unit)
            value = round(total_m2 if unit == "sq.m" else total_m2 * FT_PER_M2, 2)
            if balcony is not None:
                implied = to_m2(carpet_only, unit) + to_m2(balcony, balcony_unit)
                if implied and abs(implied - total_m2) > 0.1:
                    result.warnings.append(
                        f"{prefix_label}Unit {decoded}: the text states a total of "
                        f"{stated_total:g} {total_unit} but {carpet_only:g} carpet + "
                        f"{balcony:g} balcony comes to {implied * (1 if unit == 'sq.m' else FT_PER_M2):.2f}. "
                        f"Used the stated total — please check."
                    )
        elif value is not None and balcony is not None:
            total_m2 = to_m2(value, unit) + to_m2(balcony, balcony_unit)
            value = round(total_m2 if unit == "sq.m" else total_m2 * FT_PER_M2, 2)
        grid[decoded] = {"value": value, "unit": unit, "area_type": area_type,
                         "source": area["source"], "balcony": balcony,
                         "balcony_unit": balcony_unit, "carpet_only": carpet_only,
                         "stated_total": stated_total,
                         "non_carpet_areas": area.get("non_carpet_areas")}
    return grid, rows_used_ai, decoded_by_ai


_FALSEY = {"false", "0", "no", "n", "none", "nan", ""}


def _truthy(value):
    """
    Robust flag reading. Input files are parsed with dtype=str to preserve
    flat numbers like '0304', which turns a boolean False into the STRING
    "False" -- and bool("False") is True. Anything reading a flag off a
    DataFrame has to go through this or it silently inverts.
    """
    if isinstance(value, str):
        return value.strip().lower() not in _FALSEY
    if value is None:
        return False
    try:
        import pandas as _pd
        if _pd.isna(value):
            return True
    except Exception:
        pass
    return bool(value)


def enrich_transactions(df, aliases, result, project_dir=None):
    """
    Read every property description and resolve, per row:
        unit number, tower, floor, series position, area value + unit + TYPE

    The description is read FIRST and the structured columns only fill genuine
    gaps, because in a real export `unit_number` held "122" where the text said
    "No.6122", and the `wing` column held "5TH" (a floor). This runs on every
    row automatically -- there is nothing to press.

    The numbering convention is INFERRED, not hardcoded: each candidate rule is
    tested against the floors the descriptions state. On an Embassy Pristine
    export tower+floor+unit fitted 35/38 and floor x 100 + position fitted 2;
    on a Maharashtra export the reverse. Adding numbering.json to the project
    folder overrides the choice.

    Adds columns prefixed with "_" and returns the rule chosen.
    """
    desc_col = find_column(df.columns, "description", aliases)
    unit_col = find_column(df.columns, "unit_no", aliases)
    tower_col = find_column(df.columns, "tower", aliases)
    carpet_col = find_column(df.columns, "carpet_area", aliases)

    descriptions = df[desc_col].tolist() if desc_col else [""] * len(df)
    parsed = [extract_unit_tower(d) for d in descriptions]

    units, from_column = [], 0
    for i, info in enumerate(parsed):
        unit = info["unit_number"]
        if not unit and unit_col:
            raw = str(df[unit_col].iloc[i] or "").strip()
            digits = re.findall(r"\d{2,4}", raw)
            if digits:
                unit = digits[-1]
                from_column += 1
        units.append(unit)

    rule, note = infer_rule(
        collect_samples(descriptions, units, [p["tower_number"] for p in parsed]),
        project_dir=project_dir,
    )
    result.warnings.append(note)
    if from_column:
        result.warnings.append(
            f"{from_column} unit number(s) were not stated in the description and were "
            f"taken from the unit-number column instead."
        )

    towers, floors, positions = [], [], []
    values, area_units, area_types = [], [], []
    for i, info in enumerate(parsed):
        decoded = decode_with_rule(units[i], rule) if units[i] else None
        tower = info["tower_number"] or (decoded[0] if decoded else None)
        if not tower and tower_col:
            tower = normalize_tower(df[tower_col].iloc[i])
        stated_floor = floor_from_text(descriptions[i])
        floor = stated_floor if stated_floor is not None else (decoded[1] if decoded else None)
        position = decoded[2] if decoded else None

        value, unit_of, label = primary_area(descriptions[i])
        if value is None and carpet_col:
            raw = str(df[carpet_col].iloc[i] or "").replace(",", "").strip()
            try:
                candidate = float(raw)
                if candidate > 0:
                    value, unit_of, label = candidate, "sq.ft", "Carpet (from column)"
            except ValueError:
                pass

        towers.append(normalize_tower(tower) if tower else None)
        floors.append(floor)
        positions.append(position)
        values.append(value)
        area_units.append(unit_of)
        area_types.append(label)

    df["_unit_resolved"] = units
    df["_tower_norm"] = towers
    df["_floor_resolved"] = floors
    df["_series_resolved"] = positions
    df["_area_value"] = values
    df["_area_unit"] = area_units
    df["_area_type"] = area_types

    found = sum(1 for v in values if v is not None)
    kinds = {}
    for label in area_types:
        if label:
            kinds[label] = kinds.get(label, 0) + 1
    result.warnings.append(
        f"Read {found}/{len(df)} area(s) from the descriptions"
        + (f" — {', '.join(f'{k}: {v}' for k, v in sorted(kinds.items()))}." if kinds else ".")
    )
    return rule


def _log(result, tower_label, series_label, floor, position, issue_type, comment):
    """Record one reviewer-facing issue."""
    unit = None
    if floor is not None and position:
        unit = f"{floor * 100 + position}"
    result.issues.append({
        "tower": (tower_label or "").replace("Tower ", "") or None,
        "series": series_label, "unit": unit or series_label,
        "floor": floor, "type": issue_type, "comment": comment,
    })


def _resolve_missing_columns(df, found, result, tower_label, api_key, model,
                             use_ai, file_kind):
    """
    Fill in any column the alias table couldn't match by asking AI. Returns the
    values of `found` in their original order, with gaps filled where possible.
    """
    missing = [f for f, c in found.items() if c is None]
    if not missing or not use_ai:
        return tuple(found.values())

    resolved, notes = resolve_columns_ai(
        list(df.columns), df.head(3).to_dict("records"), missing,
        api_key=api_key, model=model, file_kind=file_kind,
    )
    prefix = f"[{tower_label}] " if tower_label else ""
    for n in notes:
        result.warnings.append(prefix + n)
    found.update(resolved)
    return tuple(found.values())


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def map_brochure_to_columns(tm, entries):
    """
    Brochure entries -> {series_col_index: entry}, plus the ones that couldn't
    be placed. Matched on the series label if the brochure names one,
    otherwise on the Unit Type labels already sitting in the template's
    "Unit Type ->" row. One brochure plan can legitimately map to SEVERAL
    series columns -- e.g. Series 1 and Series 6 are both "4 BHK Signature".
    """
    header = tm.rows[tm.section1.header_row]
    unit_row = tm.rows[tm.section1.unit_type_row]
    by_col, unmapped = {}, []
    for e in entries:
        targets = []
        if e.get("series"):
            wanted = _norm(e["series"])
            targets = [c for c in tm.section1.series_cols if _norm(header[c]) == wanted]
        if not targets and e.get("unit_type"):
            n = _norm(e["unit_type"])
            for c in tm.section1.series_cols:
                tn = _norm(unit_row[c]) if c < len(unit_row) else ""
                if tn and (tn == n or tn in n or n in tn):
                    targets.append(c)
        if not targets:
            unmapped.append(e)
            continue
        for c in targets:
            by_col.setdefault(c, e)
    return by_col, unmapped


def place_agreements(tm, agreements, result, api_key=None, use_ai_decode=False, model=None):
    """
    Agreement records -> {(floor, position): record}. Uses the plain decoding
    rule first, then AI for anything it can't handle, exactly like the
    inventory and transaction files.
    """
    known_floors = set(tm.section1.floor_rows) | set(tm.section2.floor_rows)
    placed, pending = {}, []
    for rec in agreements:
        decoded = decode_unit(rec.get("unit_no"))
        if decoded is None or decoded[0] not in known_floors:
            pending.append(rec)
            continue
        placed[decoded] = rec
        combined = rec.get("combined")
        if combined and combined["kind"] == "duplex":
            # One home spanning two floors: register it on both, so Final
            # Output can emit a single "15/16" row instead of two.
            for fl in combined["floors"]:
                if fl in known_floors:
                    placed[(fl, decoded[1])] = rec

    if pending and use_ai_decode:
        section = tm.section1
        resolved = decode_units_ai(
            [r.get("unit_no") for r in pending],
            floors=known_floors,
            num_series=section.num_series,
            api_key=api_key,
            series_labels=[tm.rows[section.header_row][c] for c in section.series_cols],
            source="agreement",
            model=model,
        )
        for rec in pending:
            hit = resolved.get(str(rec.get("unit_no")).strip())
            if hit is not None:
                placed[(hit["floor"], hit["position"])] = rec
            else:
                result.warnings.append(
                    f"'{rec.get('filename')}': unit '{rec.get('unit_no')}' could not be "
                    f"placed on the stack grid, so its area was not used."
                )
    else:
        for rec in pending:
            result.warnings.append(
                f"'{rec.get('filename')}': unit '{rec.get('unit_no')}' could not be "
                f"placed on the stack grid, so its area was not used."
            )
    return placed


def _majority_label(values, field):
    labels = [v[field] for v in values if v.get(field)]
    if not labels:
        return None
    return Counter(labels).most_common(1)[0][0]


def _process_tower_into_sheet(ws, tm: TemplateModel, txn_df: pd.DataFrame, inv_df: pd.DataFrame,
                               aliases: dict, result: ProcessingResult, api_key: str = None,
                               use_ai_fallback: bool = True, tower_label: str = "",
                               use_ai_inventory: bool = False,
                               brochure_entries=None, agreements=None,
                               tolerance_ft: float = 5.0, model: str = None):
    """Fills one worksheet for one tower's template + (already-filtered) data.
    Returns a stats dict for this tower."""

    brochure_entries = brochure_entries or []
    agreements = agreements or []
    tolerance_m = tolerance_ft / FT_PER_M2

    fill_agreement_conflict = PatternFill("solid", fgColor=C.COLOR_AGREEMENT_CONFLICT)
    fill_no_txn = PatternFill("solid", fgColor=C.COLOR_NO_TXN)
    fill_area_type_missing = PatternFill("solid", fgColor=C.COLOR_AREA_TYPE_MISSING)
    fill_area_mismatch = PatternFill("solid", fgColor=C.COLOR_AREA_MISMATCH)
    fill_header = PatternFill("solid", fgColor=C.HEADER_FILL)
    default_font = Font(name=C.DEFAULT_FONT_NAME)
    header_font = Font(name=C.DEFAULT_FONT_NAME, bold=True)

    inventory_grid, inv_decoded_by_ai, unverified_cells = build_inventory_grid(
        inv_df, aliases, result, tower_label=tower_label,
        tm=tm, api_key=api_key, use_ai_decode=use_ai_inventory, model=model,
    )
    transaction_grid, rows_used_ai, txn_decoded_by_ai = build_transaction_grid(
        txn_df, aliases, result, api_key=api_key, use_ai_fallback=use_ai_fallback,
        tower_label=tower_label, tm=tm, use_ai_decode=use_ai_inventory, model=model,
    )
    units_decoded_by_ai = inv_decoded_by_ai + txn_decoded_by_ai

    brochure_by_col, brochure_unmapped = map_brochure_to_columns(tm, brochure_entries)
    agreements_by_pos = place_agreements(
        tm, agreements, result, api_key=api_key, use_ai_decode=use_ai_inventory,
        model=model,
    )

    # Decide what KIND of area each transaction figure is, using the agreements
    # as proof: where an agreement's area matches the transaction's area, the
    # transaction's number is that agreement's type. See core/area_type.py.
    series_labels_by_pos = {
        pos: tm.rows[tm.section1.header_row][col]
        for pos, col in enumerate(tm.section1.series_cols, start=1)
    }
    resolved_types, type_warnings = confirm_area_types(
        transaction_grid, agreements_by_pos, tm.section1.series_cols,
        tolerance_m, series_of_position=series_labels_by_pos,
    )
    prefix = f"[{tower_label}] " if tower_label else ""
    for w in type_warnings:
        result.warnings.append(prefix + w)
    for e in brochure_unmapped:
        result.warnings.append(
            f"Brochure: an area for '{e.get('unit_type') or 'an unnamed unit type'}' "
            f"(p.{e.get('page')}) matched no series in this template, so it was "
            f"listed in Section 3 without a series."
        )

    by_col_values = defaultdict(list)
    for (floor, pos), data in transaction_grid.items():
        if pos - 1 < len(tm.section1.series_cols) and data["value"] is not None:
            col_idx = tm.section1.series_cols[pos - 1]
            confirmed = resolved_types.get((floor, pos)) or {}
            by_col_values[col_idx].append({
                **data, "floor": floor,
                "confirmed_area_type": confirmed.get("area_type"),
                "type_confidence": confirmed.get("confidence"),
            })

    col_majority_value = {}
    for col_idx, values in by_col_values.items():
        nums = [v["value"] for v in values if v["value"] is not None]
        if nums:
            col_majority_value[col_idx] = Counter(nums).most_common(1)[0][0]

    # 1) Pass through every template cell verbatim -- EXCEPT the cells this
    # app computes itself. Real templates are often saved from a
    # part-finished working file, so they arrive with leftover areas still
    # sitting in the stack grids. Writing those through would mix last
    # project's numbers into this run's output wherever the new data happens
    # to have no value. Everything outside the two stack grids (header
    # blocks, checklist, legend, Unit Type row, Sections 3-6) still passes
    # through completely untouched.
    stale_cells = set()
    for sec in (tm.section1, tm.section2):
        grid_cols = list(sec.series_cols)
        if sec.notes_col is not None:
            grid_cols.append(sec.notes_col)
        for row_idx in sec.floor_rows.values():
            for c in grid_cols:
                stale_cells.add((row_idx, c))
        for c in sec.series_cols:
            stale_cells.add((sec.area_type_row, c))

    # Section 3/4 rows are only cleared when we actually have something to put
    # there -- a run with no brochure leaves whatever the template already had.
    section3_rows = brochure_by_col and tm.section3 is not None
    section4_rows = agreements_by_pos and tm.section4 is not None
    n_s3 = (len(brochure_by_col) + len(brochure_unmapped)) if section3_rows else 0
    n_s4 = len(agreements_by_pos) if section4_rows else 0
    for sec, needed in ((tm.section3, n_s3), (tm.section4, n_s4)):
        if needed:
            for r in range(sec.first_data_row, sec.last_data_row + 1):
                for c in range(0, max(len(row) for row in tm.rows)):
                    stale_cells.add((r, c))

    # If a section needs more rows than the template allots, everything below
    # it shifts down rather than being overwritten.
    inserts = []
    if tm.section3 is not None and n_s3 > tm.section3.capacity:
        inserts.append((tm.section3.last_data_row, n_s3 - tm.section3.capacity))
    if tm.section4 is not None and n_s4 > tm.section4.capacity:
        inserts.append((tm.section4.last_data_row, n_s4 - tm.section4.capacity))

    def out_row(r0):
        """0-based template row -> 1-based worksheet row, allowing for any
        extra rows inserted into an earlier section."""
        return r0 + 1 + sum(extra for after, extra in inserts if r0 > after)

    for r0, row in enumerate(tm.rows):
        for c0, val in enumerate(row):
            if val and (r0, c0) not in stale_cells:
                cell = ws.cell(row=out_row(r0), column=c0 + 1, value=val)
                cell.font = default_font

    for sec in (tm.section1, tm.section2):
        for c in range(1, len(tm.rows[sec.header_row]) + 1):
            cell = ws.cell(row=out_row(sec.header_row), column=c)
            cell.font = header_font
            cell.fill = fill_header

    notes_per_floor_section1 = defaultdict(list)
    no_txn_count = 0
    area_missing_count = 0
    agreement_conflicts = 0

    # 2) Section 1 (CRE) + anomalies.
    for floor, row_idx in tm.section1.floor_rows.items():
        excel_row = out_row(row_idx)
        for pos, col_idx in enumerate(tm.section1.series_cols, start=1):
            excel_col = col_idx + 1
            series_label = tm.rows[tm.section1.header_row][col_idx]
            cell = ws.cell(row=excel_row, column=excel_col)
            cell.font = default_font

            unit_exists = (floor, pos) in inventory_grid
            txn_data = transaction_grid.get((floor, pos))

            if txn_data is None:
                # No transaction row exists for this flat at all. Only worth
                # flagging if the RERA inventory says the flat is real --
                # otherwise this cell isn't a unit in the first place.
                if unit_exists:
                    cell.value = None
                    cell.fill = fill_no_txn
                    no_txn_count += 1
                    _log(result, tower_label, series_label, floor, pos, "No CRE transaction",
                         "The builder inventory shows this unit exists but there is no "
                         "registration for it. Download the agreement or confirm the flat "
                         "number.")

            elif txn_data["value"] is None:
                # A transaction DOES exist, but no carpet area could be read
                # out of its description text (neither regex nor AI could
                # find it). This is the "area not in the Marathi text" case
                # -- pink, and it needs a human to open the agreement.
                cell.value = None
                cell.fill = fill_area_type_missing
                area_missing_count += 1
                found = txn_data.get("non_carpet_areas") or []
                if found:
                    listed = ", ".join(f"{f['value']:g} {f['unit']}" for f in found[:3])
                    _log(result, tower_label, series_label, floor, pos,
                         "Area not in Marathi text",
                         f"The description states no CARPET area — only "
                         f"{listed} (super built-up / saleable / land share), which "
                         f"measure more than carpet and cannot be used in its place. "
                         f"Download the agreement for the carpet figure.")
                else:
                    _log(result, tower_label, series_label, floor, pos,
                         "Area not in Marathi text",
                         "A registration exists but no carpet area could be read from its "
                         "description. Open the agreement to get the area.")

            else:
                value = txn_data["value"]
                unit = txn_data["unit"] or "sq.ft"
                cell.value = f"{value:g} {unit}"

                # Show the arithmetic when a balcony was folded in, so the
                # total in the cell can always be traced back to the two
                # figures the description actually stated.
                if txn_data.get("balcony") is not None:
                    joiner = ("= stated total" if txn_data.get("stated_total") is not None
                              else "=")
                    notes_per_floor_section1[excel_row].append(
                        f"{series_label}: {txn_data['carpet_only']:g} carpet + "
                        f"{txn_data['balcony']:g} balcony {joiner} {value:g} {unit}"
                    )

                # An agreement outranks the CRE transaction, but Section 1 is
                # defined as "area exactly as in the CRE Marathi text", so the
                # CRE number STAYS here and the conflict is flagged instead.
                # Final Output uses the agreement's value.
                agreement = agreements_by_pos.get((floor, pos))
                conflict = False
                if agreement is not None and agreement.get("area_value") is not None:
                    a_m = to_m2(agreement.get("area_value"), agreement.get("area_unit"))
                    c_m = to_m2(value, unit)
                    if a_m is not None and c_m is not None and abs(a_m - c_m) >= tolerance_m:
                        conflict = True
                        agreement_conflicts += 1
                        cell.fill = fill_agreement_conflict
                        _log(result, tower_label, series_label, floor, pos,
                             "Agreement vs CRE",
                             f"Agreement says {agreement.get('area_value')} "
                             f"{agreement.get('area_unit')} but the registration says "
                             f"{value:g} {unit}. The agreement is used in Final Output — "
                             f"confirm which is right.")

                if not conflict:
                    majority = col_majority_value.get(col_idx)
                    if majority is not None and value != majority:
                        cell.fill = fill_area_mismatch
                        _log(result, tower_label, series_label, floor, pos,
                             "Area differs within series",
                             f"{value:g} {unit} differs from the majority area in this "
                             f"series. Check whether this floor is genuinely different "
                             f"(refuge, setback, jodi).")

    area_type_excel_row = out_row(tm.section1.area_type_row)
    for pos_idx, col_idx in enumerate(tm.section1.series_cols, start=1):
        label, confidence = summarise_series_type(resolved_types, pos_idx)
        if label:
            ws.cell(row=area_type_excel_row, column=col_idx + 1, value=label).font = default_font
            if confidence and "agreement" not in confidence:
                result.warnings.append(
                    f"{prefix}{series_labels_by_pos[pos_idx]}: area type '{label}' is "
                    f"{confidence} — download an agreement for this series to confirm it."
                )
    # The Notes / Anomaly columns stay EMPTY on purpose: the stack grids hold
    # values and their legend colours, and every note lives in the Review
    # Tracker sheet where a reviewer can sort and close them off.

    # 3) Section 2 (MahaRERA inventory) -- ground truth, no anomaly logic.
    rera_by_col_values = defaultdict(list)
    fill_unverified = PatternFill("solid", fgColor=C.COLOR_UNVERIFIED)
    notes_per_floor_section2 = defaultdict(list)
    for floor, row_idx in tm.section2.floor_rows.items():
        excel_row = out_row(row_idx)
        for pos, col_idx in enumerate(tm.section2.series_cols, start=1):
            area = inventory_grid.get((floor, pos))
            if area is not None:
                cell = ws.cell(row=excel_row, column=col_idx + 1, value=area)
                cell.font = default_font
                verified = (floor, pos) not in unverified_cells
                if not verified:
                    # Written, not dropped -- but highlighted so it's obvious
                    # this figure was read by AI and never cross-checked.
                    cell.fill = fill_unverified
                    _log(result, tower_label,
                         tm.rows[tm.section2.header_row][col_idx], floor, pos,
                         "Unverified AI reading",
                         f"Inventory area {area} was read by AI from a scan and could not "
                         f"be cross-checked against any text. Verify before relying on it.")
                rera_by_col_values[col_idx].append(
                    {"floor": floor, "value": area, "verified": verified}
                )

    area_type_excel_row_s2 = out_row(tm.section2.area_type_row)
    for pos_idx, col_idx in enumerate(tm.section2.series_cols, start=1):
        col_has_data = any(
            inventory_grid.get((floor, pos_idx)) is not None for floor in tm.section2.floor_rows
        )
        if col_has_data:
            ws.cell(row=area_type_excel_row_s2, column=col_idx + 1, value="sq.m").font = default_font

    # 4) Tally row for Section 1.
    if tm.section1.tally_row is not None:
        no_txn_col = tm.tally_value_col(tm.section1, "No-txn")
        area_missing_col = tm.tally_value_col(tm.section1, "Area missing")
        tally_excel_row = out_row(tm.section1.tally_row)
        if no_txn_col is not None:
            ws.cell(row=tally_excel_row, column=no_txn_col + 1, value=no_txn_count).font = default_font
        if area_missing_col is not None:
            ws.cell(row=tally_excel_row, column=area_missing_col + 1, value=area_missing_count).font = default_font

    _paint_legend(ws, tm, out_row, default_font)

    # ---- Section 3: builder brochure, one row per series ----
    if section3_rows:
        s3 = tm.section3
        # Rows are written to CONSECUTIVE physical rows starting at the
        # section's first data row. out_row() must not be applied per row here
        # -- overflow rows are the ones the shift makes room for, so shifting
        # them again would scatter them past the next section.
        r = out_row(s3.first_data_row)
        for col_idx in tm.section1.series_cols:
            e = brochure_by_col.get(col_idx)
            if e is None:
                continue
            _write_s3_row(ws, r, s3, default_font,
                          tm.rows[tm.section1.header_row][col_idx],
                          tm.rows[tm.section1.unit_type_row][col_idx]
                          if col_idx < len(tm.rows[tm.section1.unit_type_row]) else None, e)
            r += 1
        for e in brochure_unmapped:
            _write_s3_row(ws, r, s3, default_font, None, e.get("unit_type"), e)
            r += 1

    # ---- Section 4: agreements, one row per uploaded file ----
    if section4_rows:
        s4 = tm.section4
        r = out_row(s4.first_data_row)
        for n, ((floor, pos), rec) in enumerate(
                sorted(agreements_by_pos.items(), key=lambda kv: (kv[0][0], kv[0][1])), start=1):
            series_label = (tm.rows[tm.section1.header_row][tm.section1.series_cols[pos - 1]]
                            if pos - 1 < len(tm.section1.series_cols) else None)
            _write_s4_row(ws, r, s4, default_font, n, floor, series_label, rec)
            r += 1

    max_col = max((len(r) for r in tm.rows), default=12)
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    return {
        "no_txn": no_txn_count,
        "area_missing": area_missing_count,
        "rows_used_ai": rows_used_ai,
        "units_decoded_by_ai": units_decoded_by_ai,
        "agreement_conflicts": agreement_conflicts,
        "by_col_values": dict(by_col_values),
        "rera_by_col_values": dict(rera_by_col_values),
        "agreements_by_pos": agreements_by_pos,
        "brochure_by_col": brochure_by_col,
    }


def _paint_legend(ws, tm, out_row, font):
    """
    Fill in the LEGEND block's colour swatches, and add any colour this app
    uses that the printed legend doesn't list.

    A template exported as CSV has legend labels but no fills, so the reader
    is told a colour exists without being shown it. The swatch is the cell
    immediately left of each label; it is left alone if it already holds text.
    """
    legend_start = None
    for i, row in enumerate(tm.rows):
        if row and row[0] and row[0].strip().upper().startswith("LEGEND"):
            legend_start = i
            break
    if legend_start is None:
        return

    # The block runs until the next section marker.
    legend_end = len(tm.rows)
    for i in range(legend_start + 1, len(tm.rows)):
        row = tm.rows[i]
        if row and row[0] and "section" in row[0].lower():
            legend_end = i
            break

    free_slots = []   # (template_row, swatch_col, label_col) with room to spare
    for i in range(legend_start + 1, legend_end):
        row = tm.rows[i]
        if not row:
            continue
        painted_here = False
        for c, cell in enumerate(row):
            text = (cell or "").strip()
            if not text or c == 0:
                continue
            color = C.LEGEND_COLORS.get(text.lower())
            if color and not (row[c - 1] or "").strip():
                sw = ws.cell(row=out_row(i), column=c)      # 1-based == c-1 + 1
                sw.fill = PatternFill("solid", fgColor=color)
                sw.border = _thin_border()
                painted_here = True
        # A row whose right-hand pair is entirely empty can host a new entry.
        if painted_here and len(row) > 5 and not (row[4] or "").strip() \
                and not (row[5] or "").strip():
            free_slots.append((i, 4, 5))

    for (label, color) in C.EXTRA_LEGEND_ENTRIES:
        if not free_slots:
            break
        i, swatch_col, label_col = free_slots.pop(0)
        sw = ws.cell(row=out_row(i), column=swatch_col + 1)
        sw.fill = PatternFill("solid", fgColor=color)
        sw.border = _thin_border()
        ws.cell(row=out_row(i), column=label_col + 1, value=label).font = font


def _thin_border():
    from openpyxl.styles import Border, Side
    side = Side(style="thin", color="FF808080")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_s3_row(ws, excel_row, s3, font, series_label, unit_type, e):
    """One Section 3 (brochure) row, written by column label so a template
    with a different column order still works."""
    note = f"Brochure p.{e.get('page')}"
    if e.get("verified") is False:
        note += " (read from page image, unverified)"
    for labels, value in [
        (("series",), series_label),
        (("unit type",), e.get("unit_type") or unit_type),
        (("applicable floors",), e.get("applicable_floors")),
        (("area type",), e.get("area_type")),
        (("carpet area",), f"{e.get('carpet_area')} {e.get('carpet_unit')}"
            if e.get("carpet_area") is not None else None),
        (("balcony",), e.get("balcony_area")),
        (("notes",), note),
    ]:
        c = s3.col(*labels)
        if c is not None and value is not None:
            ws.cell(row=excel_row, column=c + 1, value=value).font = font


def _write_s4_row(ws, excel_row, s4, font, n, floor, series_label, rec):
    """One Section 4 (agreement) row."""
    is_photo = rec.get("is_photo")
    note = rec.get("filename") or ""
    if rec.get("page"):
        note += f" p.{rec['page']}"
    if rec.get("agreement_date"):
        note += f" · dated {rec['agreement_date']}"
    if is_photo:
        # A photo has no text layer, so the model's own transcription of the
        # phrase it read is what a reviewer checks against the image.
        note += " · photo, unverified"
        if rec.get("source_text"):
            note += f' · reads: "{rec["source_text"]}"'
    elif rec.get("verified") is False:
        note += " · read from page image, unverified"

    ws.cell(row=excel_row, column=1, value=n).font = font
    for labels, value in [
        (("unit no",), rec.get("unit_no")),
        (("series",), series_label),
        (("floor",), floor),
        (("reason for download",), "Agreement supplied as evidence"),
        (("loaded",), "Yes"),
        (("area type in agreement", "area type"), rec.get("area_type")),
        # A balcony-only photo has no carpet figure; show the balcony value it
        # DOES have rather than printing "None sq.ft".
        (("area value exact", "area value"),
         f"{rec.get('area_value')} {rec.get('area_unit')}"
         if rec.get("area_value") is not None
         else (f"{rec.get('balcony_area')} {rec.get('area_unit')} (balcony)"
               if rec.get("balcony_area") is not None else None)),
        # The photo IS the area-text screenshot this column asks for.
        (("area text screenshot",), rec.get("filename") if is_photo else None),
        (("notes",), note),
    ]:
        c = s4.col(*labels)
        if c is not None and value is not None:
            ws.cell(row=excel_row, column=c + 1, value=value).font = font


def _agreements_for_tower(agreements, letter, single_tower, result):
    """
    Agreements belonging to this tower. An agreement that doesn't state a
    tower is assumed to belong to the only tower there is; with several
    towers it can't be placed safely, so it's skipped with a warning rather
    than guessed onto the wrong sheet.
    """
    mine = []
    for a in agreements:
        tower = a.get("tower")
        if tower == letter or tower is None:
            mine.append(a)
        elif single_tower:
            # The document's tower label rarely matches the sheet's: a wing
            # letter read off an agreement ("B") won't equal a tower value
            # derived from an inventory file ("SKYVISTASBLUEZ"). With only one
            # tower there is nowhere else the flat could go, so use it and
            # mention the mismatch -- dropping it was silently losing evidence.
            mine.append(a)
            result.warnings.append(
                f"'{a.get('filename')}': the document says tower '{tower}' but this sheet "
                f"is '{letter}'. There's only one tower, so it was applied here anyway — "
                f"check it's the right one."
            )
        else:
            result.warnings.append(
                f"'{a.get('filename')}': says tower '{tower}', which matches none of this "
                f"project's towers, so it was NOT applied. Rename the file or check the "
                f"tower letter."
            )
    return mine


def _find_tower_dirs(project_dir: Path):
    """Subfolders of project_dir that contain their own template.csv."""
    if not project_dir.is_dir():
        return []
    return sorted(
        [p for p in project_dir.iterdir() if p.is_dir() and (p / "template.csv").exists()],
        key=lambda p: p.name,
    )


def generate_stack_view(project_dir: str, transaction_csv_path: str, inventory_path: str,
                         output_path: str, api_key: str = None, use_ai_fallback: bool = True,
                         use_ai_inventory: bool = False, brochure_path: str = None,
                         agreement_files=None, tolerance_ft: float = 5.0,
                         use_ai_documents: bool = True,
                         model: str = None) -> ProcessingResult:
    # An unspecified key or model comes from config.json / .env, so callers
    # and scripts don't need to know where those live.
    settings = load_settings()
    api_key = api_key or settings["api_key"]
    model = model or settings["model"]

    result = ProcessingResult()
    project_dir = Path(project_dir)

    project_aliases = load_aliases(project_dir)

    raw_txn_df = pd.read_csv(transaction_csv_path, dtype=str, keep_default_na=False)
    txn_df = clean_transaction_df(raw_txn_df, project_aliases)

    inv_ext = Path(inventory_path).suffix.lower()
    if inv_ext == ".pdf":
        try:
            raw_inv_df, pdf_warnings = extract_inventory_from_pdf(
                inventory_path, api_key=api_key, use_ai=use_ai_inventory, model=model
            )
            result.warnings.extend(pdf_warnings)
        except Exception as e:
            result.warnings.append(
                f"Unexpected error reading the inventory PDF ({type(e).__name__}: {e}). "
                f"Section 2 will be left empty."
            )
            raw_inv_df = pd.DataFrame(columns=["tower", "flat_no", "carpet_area"])
    else:
        raw_inv_df = pd.read_excel(inventory_path, dtype=str)
    inv_df = clean_inventory_df(raw_inv_df)

    # Evidence documents are read once, then split per tower by their own
    # stated tower letter.
    # A brochure states its areas inside floor-plan artwork and an agreement
    # buries them in legalese, so neither can be read without AI -- there is no
    # deterministic fallback to degrade to. If the switch is off, say so
    # plainly rather than silently producing empty Sections 3 and 4.
    brochure_entries, agreements = [], []
    if not use_ai_documents:
        if brochure_path or agreement_files:
            result.warnings.append(
                "A brochure/agreement was supplied but reading them with AI is switched "
                "off, so Sections 3 and 4 were left as they are in the template."
            )
    else:
        if brochure_path:
            brochure_entries, bw = extract_brochure(brochure_path, api_key=api_key, model=model)
            result.warnings.extend(bw)
            result.brochure_areas_found = len(brochure_entries)
        if agreement_files:
            agreements, aw = extract_agreements(agreement_files, api_key=api_key, model=model)
            result.warnings.extend(aw)
            result.agreements_read = len(agreements)

    # Read every description now: unit, tower, floor, series and area, with the
    # numbering convention inferred from the data rather than assumed.
    numbering_rule = enrich_transactions(txn_df, project_aliases, result,
                                         project_dir=project_dir)

    tower_dirs = _find_tower_dirs(project_dir)          # {LETTER: Path}, only towers with their OWN template
    tower_dir_by_letter = {d.name.upper(): d for d in tower_dirs}
    root_template = project_dir / "template.csv"
    has_root_template = root_template.exists()

    # Does the data itself carry a Tower/Wing column with actual values?
    tower_col_txn = find_column(txn_df.columns, "tower", project_aliases)
    tower_col_inv = find_column(inv_df.columns, "tower", project_aliases)

    # enrich_transactions() already resolved the tower per row, reading the
    # description first (e.g. "TOWER-4(DAFFODIL)") and only then any tower
    # column -- so it must not be recomputed from the column here.
    if "_tower_norm" not in txn_df.columns:
        txn_df["_tower_norm"] = None
    txn_has_tower = txn_df["_tower_norm"].notna().any()

    if tower_col_inv:
        inv_df["_tower_norm"] = inv_df[tower_col_inv].apply(normalize_tower)
    else:
        inv_df["_tower_norm"] = None

    data_letters = set(txn_df["_tower_norm"].dropna()) | set(inv_df["_tower_norm"].dropna())

    # Split into per-tower sheets if EITHER: tower-specific template folders
    # exist, OR the data itself actually contains tower values to split on.
    should_split = bool(tower_dir_by_letter) or bool(data_letters)

    wb = Workbook()
    default_sheet = wb.active
    tower_entries = []  # feeds the Final Output (NocoDB hand-off) sheet, built after all towers are processed

    if not should_split:
        # ---- Simple single-tower project: exactly one sheet, no splitting. ----
        if not has_root_template:
            raise ValueError(
                f"No template.csv found directly under '{project_dir}', and no "
                f"tower subfolders (each with their own template.csv) either."
            )
        tm = TemplateModel.from_csv(str(root_template))
        default_sheet.title = "Stack View"
        stats = _process_tower_into_sheet(
            default_sheet, tm, txn_df, inv_df, project_aliases, result,
            api_key=api_key, use_ai_fallback=use_ai_fallback,
            use_ai_inventory=use_ai_inventory,
            brochure_entries=brochure_entries, agreements=agreements,
            tolerance_ft=tolerance_ft, model=model,
        )
        result.no_txn_count = stats["no_txn"]
        result.area_missing_count = stats["area_missing"]
        result.rows_used_ai = stats["rows_used_ai"]
        result.units_decoded_by_ai = stats["units_decoded_by_ai"]
        result.agreement_conflicts = stats["agreement_conflicts"]
        tower_entries.append({
            "letter": "", "tm": tm,
            "by_col_values": stats["by_col_values"],
            "rera_by_col_values": stats["rera_by_col_values"],
            "agreements_by_pos": stats["agreements_by_pos"],
            "brochure_by_col": stats["brochure_by_col"],
        })

    else:
        # ---- Multi-tower: one sheet per tower found in the data. ----
        # Each tower uses its own template if one exists for it, otherwise
        # falls back to the shared template at the project root.
        wb.remove(default_sheet)

        if not tower_col_txn:
            result.warnings.append(
                f"Transaction file: no Tower/Wing column found (looked for: "
                f"{project_aliases.get('tower')}). Every tower's Section 1 will be empty."
            )
        if not tower_col_inv:
            result.warnings.append(
                f"Inventory file: no Tower/Wing column found (looked for: "
                f"{project_aliases.get('tower')}). Every tower's Section 2 will be empty."
            )

        all_letters = sorted(set(tower_dir_by_letter.keys()) | data_letters)
        skipped_letters = []

        for letter in all_letters:
            tower_dir = tower_dir_by_letter.get(letter)
            if tower_dir is not None:
                template_path = tower_dir / "template.csv"
                tower_aliases = load_aliases(tower_dir)
                for field, names in project_aliases.items():
                    tower_aliases.setdefault(field, [])
                    tower_aliases[field] = tower_aliases[field] + [n for n in names if n not in tower_aliases[field]]
            elif has_root_template:
                template_path = root_template
                tower_aliases = project_aliases
            else:
                skipped_letters.append(letter)
                continue

            filtered_txn = (txn_df[txn_df["_tower_norm"] == letter]
                            if txn_has_tower else txn_df.iloc[0:0])
            filtered_inv = inv_df[inv_df["_tower_norm"] == letter] if tower_col_inv else inv_df.iloc[0:0]

            tm = TemplateModel.from_csv(str(template_path))
            ws = wb.create_sheet(title=f"Tower {letter}"[:31])  # Excel sheet-name length limit

            stats = _process_tower_into_sheet(
                ws, tm, filtered_txn, filtered_inv, tower_aliases, result,
                api_key=api_key, use_ai_fallback=use_ai_fallback, tower_label=f"Tower {letter}",
                use_ai_inventory=use_ai_inventory,
                brochure_entries=brochure_entries,
                agreements=_agreements_for_tower(
                    agreements, letter, single_tower=len(all_letters) == 1, result=result
                ),
                tolerance_ft=tolerance_ft, model=model,
            )
            result.per_tower[letter] = stats
            result.no_txn_count += stats["no_txn"]
            result.area_missing_count += stats["area_missing"]
            result.rows_used_ai += stats["rows_used_ai"]
            result.units_decoded_by_ai += stats["units_decoded_by_ai"]
            result.agreement_conflicts += stats["agreement_conflicts"]
            tower_entries.append({
                "letter": letter, "tm": tm,
                "by_col_values": stats["by_col_values"],
                "rera_by_col_values": stats["rera_by_col_values"],
                "agreements_by_pos": stats["agreements_by_pos"],
                "brochure_by_col": stats["brochure_by_col"],
            })

        if skipped_letters:
            result.warnings.append(
                f"Tower(s) {', '.join(skipped_letters)} appear in your data but have no "
                f"tower-specific template AND no default project-level template.csv to fall "
                f"back on -- their rows were skipped. Add a template for them to include them."
            )

    fallback_society_name = project_dir.name.replace("_", " ")
    _, final_rows = build_final_output_sheet(
        wb, tower_entries, fallback_society_name, tolerance_ft=tolerance_ft
    )

    # Rows the Final Output flagged become tracker issues too, so a reviewer
    # sees stack-level and hand-off-level problems in one list.
    for row in final_rows:
        status = row.get("status") or ""
        if status.startswith("Needs Review"):
            issue_type = ("CRE vs RERA" if "differ" in status else "Area type unconfirmed")
            if "brochure only" in status:
                issue_type = "Brochure only"
            elif "unverified" in status:
                issue_type = "Unverified AI reading"
            result.issues.append({
                "tower": row.get("tower_letter") or None,
                "series": row.get("series"), "unit": row.get("series"),
                "floor": (row.get("floor_list") or [None])[0],
                "type": issue_type,
                "comment": f"Floors {row.get('floors')} at {row.get('area')}: "
                           f"{status.replace('Needs Review — ', '')}",
            })

    # Which agreements would settle what's still unconfirmed.
    requests = suggest_agreements(final_rows, tower_entries, tolerance_ft=tolerance_ft)
    if requests and use_ai_documents:
        drafted = draft_reasons_with_ai(requests, api_key=api_key, model=model)
        for req, reason in zip(requests, drafted):
            if reason:
                req["reason"] = reason
    result.agreement_requests = len(requests)
    for req in requests:
        result.issues.append({
            "tower": req.get("tower"), "series": req.get("series"),
            "unit": req.get("unit"), "floor": req.get("floor"),
            "type": "Area type unconfirmed",
            "comment": f"No agreement evidence for floors {req.get('floors')} "
                       f"({req.get('area')}). Suggested download: unit {req.get('unit')}.",
        })

    build_agreement_requests_sheet(wb, requests, fallback_society_name)
    build_review_tracker_sheet(wb, result.issues, fallback_society_name)

    wb.save(output_path)
    result.output_path = output_path
    return result
